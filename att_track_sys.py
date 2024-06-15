import os
import openpyxl
import base64
import json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.auth.transport.requests import Request  # Import Request from google.auth.transport.requests

# Path to the client_id.json file
CLIENT_SECRET_FILE = 'client_id.json'
SCOPES = ['https://www.googleapis.com/auth/gmail.send']

# Loading the excel sheet
file_path = 'E:\\Python Projects\\Attendance Tracker sys using mail and excel\\attendance.xlsx'
book = openpyxl.load_workbook(file_path)

# Choose the sheet
sheet = book['Sheet1']

# Counting number of rows / students
r = sheet.max_row

# Variable for looping for input
resp = 1

# List of students to remind
l1 = []

# To concatenate list of roll numbers with lack of attendance
l2 = ""

# List of roll numbers with lack of attendance
l3 = []

# Staff mail ids
staff_mails = ['anil687305@gmail.com']

# Warning messages
m1 = '''Dear Student,

I hope this message finds you well. I am writing to bring to your attention that according to our records, you are approaching the maximum allowable absences for the DBMS (Database Management System) class.

Attendance plays a crucial role in your academic progress and success. Consistent presence in class ensures that you fully benefit from lectures, discussions, and hands-on exercises that are essential for mastering the concepts covered in the course.

As you approach the attendance limit, it is imperative that you prioritize attending all future sessions of the DBMS class. Your commitment to regular attendance will not only enhance your understanding but also demonstrate your dedication to academic excellence.

Please take this as a gentle reminder to review your attendance and make necessary adjustments to ensure compliance with the attendance policy of our institution. If you anticipate any challenges in attending upcoming sessions, I encourage you to reach out to me or your course instructor at your earliest convenience.

Thank you for your attention to this matter. We are here to support you in achieving your academic goals, and we look forward to your continued engagement and participation in the DBMS class.

Best regards,'''
m2 = '''Dear Student,

I hope this message finds you well. I am writing to bring to your attention that according to our records, you are approaching the maximum allowable absences for the DS (Data Structures) class.

Attendance plays a crucial role in your academic progress and success. Consistent presence in class ensures that you fully benefit from lectures, discussions, and hands-on exercises that are essential for mastering the concepts covered in the course.

As you approach the attendance limit, it is imperative that you prioritize attending all future sessions of the DS class. Your commitment to regular attendance will not only enhance your understanding but also demonstrate your dedication to academic excellence.

Please take this as a gentle reminder to review your attendance and make necessary adjustments to ensure compliance with the attendance policy of our institution. If you anticipate any challenges in attending upcoming sessions, I encourage you to reach out to me or your course instructor at your earliest convenience.

Thank you for your attention to this matter. We are here to support you in achieving your academic goals, and we look forward to your continued engagement and participation in the DS class.

Best regards,'''
m3 = '''Dear Student,

I hope this message finds you well. I am writing to bring to your attention that according to our records, you are approaching the maximum allowable absences for the Python Class.

Attendance plays a crucial role in your academic progress and success. Consistent presence in class ensures that you fully benefit from lectures, discussions, and hands-on exercises that are essential for mastering the concepts covered in the course.

As you approach the attendance limit, it is imperative that you prioritize attending all future sessions of the Python class. Your commitment to regular attendance will not only enhance your understanding but also demonstrate your dedication to academic excellence.

Please take this as a gentle reminder to review your attendance and make necessary adjustments to ensure compliance with the attendance policy of our institution. If you anticipate any challenges in attending upcoming sessions, I encourage you to reach out to me or your course instructor at your earliest convenience.

Thank you for your attention to this matter. We are here to support you in achieving your academic goals, and we look forward to your continued engagement and participation in the Python class.

Best regards,'''

def savefile():
    book.save(file_path)
    print("Saved!")

def get_service():
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    service = build('gmail', 'v1', credentials=creds)
    return service

def send_email(to, subject, body):
    try:
        service = get_service()
        message = MIMEMultipart()
        message['to'] = to
        message['subject'] = subject
        message.attach(MIMEText(body, 'plain'))
        raw = base64.urlsafe_b64encode(message.as_string().encode()).decode()
        message = {
            'raw': raw
        }
        service.users().messages().send(userId="me", body=message).execute()
        print(f"Email sent to {to}")
    except HttpError as error:
        print(f"An error occurred: {error}")
        print(f"Error details: {error.resp.status} {error._get_reason()}")

def check(no_of_days, row_num, b):
    global staff_mails, l2, l3
    for student in range(len(row_num)):
        if no_of_days[student] == 2:
            if b == 1:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                send_email(l1[-1], 'Attendance Warning', m1)
            elif b == 2:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                send_email(l1[-1], 'Attendance Warning', m2)
            else:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                send_email(l1[-1], 'Attendance Warning', m3)
        elif no_of_days[student] > 2:
            if b == 1:
                l2 += str(sheet.cell(row=row_num[student], column=1).value) + " "
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "DBMS"
            elif b == 2:
                l2 += str(sheet.cell(row=row_num[student], column=1).value) + " "
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "DS"
            else:
                l2 += str(sheet.cell(row=row_num[student], column=1).value) + " "
                l3.append(sheet.cell(row=row_num[student], column=2).value)
                subject = "PYTHON"
        if l2 and l3:
            msg1 = "You have lack of attendance in " + subject + "!!!"
            msg2 = "The following students have lack of attendance in your subject: " + l2
            for email in l3:
                send_email(email, 'Attendance Alert', msg1)
            if b <= len(staff_mails):
                send_email(staff_mails[b - 1], 'Lack of Attendance Report', msg2)
            else:
                print("Error: Invalid subject selection or no staff email configured.")

while resp == 1:
    print("1--->DBMS\n2--->DS\n3--->PYTHON")
    y = int(input("Enter subject: "))
    
    # Validate subject selection
    if y not in [1, 2, 3]:
        print("Invalid subject selection. Please choose between 1 (DBMS), 2 (DS), or 3 (PYTHON).")
        continue
    
    no_of_absentees = int(input('No. of absentees: '))
    x = []
    for _ in range(no_of_absentees):
        x.append(int(input('Roll no: ')))
    
    row_num = []
    no_of_days = []

    for student in x:
        found = False
        for i in range(2, r + 1):
            if sheet.cell(row=i, column=1).value == student:
                found = True
                if y == 1:
                    m = sheet.cell(row=i, column=3).value
                    m += 1
                    sheet.cell(row=i, column=3).value = m
                    no_of_days.append(m)
                    row_num.append(i)
                elif y == 2:
                    m = sheet.cell(row=i, column=4).value
                    m += 1
                    sheet.cell(row=i, column=4).value = m
                    no_of_days.append(m)
                    row_num.append(i)
                elif y == 3:
                    m = sheet.cell(row=i, column=5).value
                    m += 1
                    sheet.cell(row=i, column=5).value = m
                    no_of_days.append(m)
                    row_num.append(i)
        if not found:
            print(f"Roll number {student} not found in the sheet. Please check and retry.")
    
    # Call check function only if subject selection is valid
    if y in [1, 2, 3]:
        check(no_of_days, row_num, y)
    else:
        print("Skipping attendance check due to invalid subject selection.")
    
    resp = int(input('Another subject? 1---->yes 0--->no: '))

print("Process completed. Exiting program.")


savefile()
